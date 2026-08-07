from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# Load workbook
wb = load_workbook('Grades.xlsx')

# Select the sheet from the work book
ws = wb.active

# Get cell value
print(ws['A1'].value)

# Modify cell value and save
ws['A2'] = "Tommy"
wb.save("Grades.xlsx")

# Accessing Multiple Cells
for row in range(1, 3):
    for col in range(1, 5):
        char = get_column_letter(col)
        print(ws[char + str(row)].value)
        
# Merging Cells
# ws.merge_cells("A1:D2")

# Inserting and Deleting rows
# ws.insert_rows(7)